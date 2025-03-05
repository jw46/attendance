import os
import openpyxl
import app.apputil.util as util
import app.data.app_data as app_data
import app.data.class_loader as cl
import app.data.student_loader as sl
import app.apputil.config as config
import app.data.loader as loader

def load_template():
	return openpyxl.load_workbook(filename=util.get_spreadsheet_folder() + app_data.selected_group + '.xlsx')

def put_class_titles_column(column, row, sheet, date_str, title):
	while sheet.cell(row=row, column=column).value is not None:
		if util.date_to_string(sheet.cell(row=row, column=column).value) == date_str and sheet.cell(row=row, column=column + 1).value is None:
			sheet.cell(row=row, column=column + 1).value = title
			return False
		row = row + 1
	return True

def put_class_titles(sheet, date_str, title):
	if put_class_titles_column(2, 3, sheet, date_str, title):
		return put_class_titles_column(2, 3, sheet, date_str, title)
	return False

def add_class_titles(workbook):
	class_df = cl.load()
	sheet = workbook[workbook.sheetnames[1]]
	for index, row in class_df.iterrows():
		put_class_titles(sheet, row['Class Dates'], row['Class Title1'])
		put_class_titles(sheet, row['Class Dates'], row['Class Title2'])
	return workbook


def add_students(workbook):
	students_df = sl.load()
	class_df = cl.load()
	ss_row = 5
	ss_column = 2
	sheet = workbook[workbook.sheetnames[2]]
	for index, row in students_df.iterrows():
		sheet.cell(row=ss_row, column=ss_column).value = row[config.STUDENT_NAME_COLUMN_NAME]
		ss_row = ss_row + 1
	return workbook

def get_att_files():
	att_files = os.listdir(util.get_att_folder())
	att_files = list(filter(lambda x: x.startswith(app_data.selected_group), att_files))
	return att_files

def fill_column(ss_column, sheet, att_df):
	row = 5
	name_column = 2
	while sheet.cell(row=row, column=name_column).value is not None:
		name = sheet.cell(row=row, column=name_column).value
		result_df = att_df.query(f'{config.STUDENT_NAME_COLUMN_NAME} == "{name}"')
		if len(result_df) > 0:
			att = result_df['attendance'].values[0]
			if att:
				sheet.cell(row=row, column=ss_column).value = att
		row = row + 1

def match_column(ss_column, ss_column_date, sheet, att_files):
	filered_att_files = list(filter(lambda x: x[8:13].replace('-','\n') == ss_column_date, att_files))
	if len(filered_att_files) == 0:
		return False
	att_df = loader.open_df(util.get_att_folder() + filered_att_files[0])
	fill_column(ss_column, sheet, att_df)


def add_attendance(workbook):
	sheet = workbook[workbook.sheetnames[2]]
	att_files = get_att_files()
	ss_row = 4
	ss_column = 6
	while sheet.cell(row=ss_row, column=ss_column).value is not None:
		match_column(ss_column, sheet.cell(row=ss_row, column=ss_column).value, sheet, att_files)
		ss_column = ss_column + 1
	return workbook

def save_spreadsheet(workbook):
	filepath = util.get_temp_folder() + app_data.selected_group + '.xlsx'
	workbook.save(filename=filepath)
	return filepath

def share_sreadsheet():
	pass

def export():
	workbook = load_template()
	add_class_titles(workbook)
	add_students(workbook)
	add_attendance(workbook)
	filepath = save_spreadsheet(workbook)
