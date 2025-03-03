import os
import openpyxl
# import dialogs
import pandas as pd
from pathlib import Path
import app.apputil.config as config
import app.apputil.util as util
import app.data.groups_loader as groups_loader
import app.data.data_saver as data_saver

def get_date_values(column, sheet):
	output = []
	row = 3
	while sheet.cell(row=row, column=column).value is not None:
		output.append(sheet.cell(row=row, column=column).value)
		row = row + 1
	return output

def get_class_dates(workbook):
	sheet = workbook[workbook.sheetnames[1]]
	class_dates = get_date_values(2, sheet) + get_date_values(7, sheet)
	return sorted(set(class_dates))

def save_if_not_exist(dates, excel):
	csv_path = util.get_classes_folder() + excel.replace('.xlsx','.csv')
	if os.path.exists(csv_path):
		return csv_path + ' - file already there, nothing done.'
	else:
		df = pd.DataFrame(dates, columns=[config.CLASS_DATES_COLUMN_NAME])
		df[config.CLASS_TITLE_COLUMN_NAME1] = pd.Series(dtype='str')
		df[config.CLASS_TITLE_COLUMN_NAME2] = pd.Series(dtype='str')
		ds = data_saver.DataSaver()
		ds.save(df, csv_path)
		return csv_path + ' - saved to app data.'
		
		

def run():
	excel_folder = util.get_spreadsheet_folder()
	excel_list = os.listdir(excel_folder)
	for excel in excel_list:
		workbook = openpyxl.load_workbook(filename=util.get_spreadsheet_folder() + excel)
		dates = get_class_dates(workbook)
		print(save_if_not_exist(dates, excel))

def save_excel():
	WORKBOOK.save(filename=EXCEL_FILEPATH)

#def share():
#	dialogs.share_url(Path(EXCEL_FILEPATH).absolute().resolve().as_uri())
	
#add_class_title("the good lession")
#share()
	
	
	
