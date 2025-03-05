import pytest
import openpyxl
import app.apputil.export_to_spreadsheet as xts
import app.data.app_data as app_data
import app.apputil.util as util
import app.data.loader as loader


def save_spreadsheet(workbook):
    print('save workbook')
    workbook.save('/home/jon/agh2/spreadsheets_test/' + app_data.selected_group + '_.xlsx')

# def test_load_template():
#     app_data.selected_group = '104'
#     result = xts.load_template()
#     print(result)

# def test_add_class_titles():
#     app_data.selected_group = '104'
#     workbook = xts.load_template()
#     result = xts.add_class_titles(workbook)
#     print(result)
#     save_spreadsheet(result)

# def test_add_students():
#     app_data.selected_group = '104'
#     workbook = xts.load_template()
#     result = xts.add_students(workbook)
#     print(result)
#     save_spreadsheet(result)

# def test_get_att_files():
#     app_data.selected_group = '104'
#     result = xts.get_att_files()
#     print(repr(result))

# def test_fill_column():
#     ss_column = 6
#     workbook = openpyxl.load_workbook(filename='/home/jon/agh2/spreadsheets_test/104.xlsx')
#     sheet = workbook[workbook.sheetnames[2]]
#     att_df = att_df = loader.open_df(util.get_att_folder() + '1042025-03-03.csv')
#     result = xts.fill_column(ss_column, sheet, att_df)

# def test_match_column():
#     result = xts.match_column(1, '03\n03', None, ['1042025-03-03.csv'])
#     print(result)

# def test_add_attendance():
#     app_data.selected_group = '104'
#     workbook = openpyxl.load_workbook(filename='/home/jon/agh2/spreadsheets_test/104.xlsx')
#     result = xts.add_attendance(workbook)
#     print(repr(result))
#     save_spreadsheet(result)

def test_export():
    app_data.selected_group = '104'
    print(xts.export())

def test_share_sreadsheet():
    pass